from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])  # 셀 값의 객체를 반환
print(ws["A2"].value)  # 셀 값 자체를 반환
print(ws["A10"].value)  # 값이 없을 땐 None을 출력

print(ws.cell(row=1, column=1).value)  # A1셀을 지정하고 있음.
print(ws.cell(row=1, column=2).value)
print(ws.cell(column=2, row=3).value)

c = ws.cell(column=3, row = 1, value=10) # 값 바로쓰기
print(c.value)

from random import *

index = 0
for x in range(1,11):
    for y in range(1,11):
        #ws.cell(row=x, column=y, value=randint(0,100))
        ws.cell(row=x, column=y, value=index)
        index +=1
wb.save("sample.xlsx")
