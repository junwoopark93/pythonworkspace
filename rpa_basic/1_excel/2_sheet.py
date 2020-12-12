from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet() # 새로운 sheet 생성
ws.title = "MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "dbaceb" #RGB형태로 값을 넣어주면 됨

ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 시트 생성
ws2 = wb.create_sheet("NewSheet", 1) # 인덱스 값을 지정할 수 있음.

new_ws = wb["NewSheet"] #Dict 형태로 Sheet에 접근 가능.

print(wb.sheetnames) # 모든 sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"
print(wb.sheetnames) # 모든 sheet 이름 확인


wb.save("sample.xlsx")